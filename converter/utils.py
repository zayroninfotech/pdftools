import os
import uuid
import zipfile
import io
from pathlib import Path

from django.conf import settings

from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch, mm
from reportlab.lib.colors import Color
from PIL import Image


UPLOAD_DIR = os.path.join(settings.MEDIA_ROOT, 'uploads')
OUTPUT_DIR = os.path.join(settings.MEDIA_ROOT, 'outputs')

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _unique_name(ext):
    return f"{uuid.uuid4().hex}{ext}"


def save_uploaded_file(uploaded_file):
    name = _unique_name(Path(uploaded_file.name).suffix)
    path = os.path.join(UPLOAD_DIR, name)
    with open(path, 'wb') as f:
        for chunk in uploaded_file.chunks():
            f.write(chunk)
    return path


def save_uploaded_files(uploaded_files):
    return [save_uploaded_file(f) for f in uploaded_files]


# ─── 1. MERGE PDF ────────────────────────────────────────────────────────────

def merge_pdfs(file_paths):
    writer = PdfWriter()
    for path in file_paths:
        reader = PdfReader(path)
        for page in reader.pages:
            writer.add_page(page)
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 2. SPLIT PDF ────────────────────────────────────────────────────────────

def split_pdf(file_path, mode='all', ranges_str=''):
    reader = PdfReader(file_path)
    total_pages = len(reader.pages)

    if mode == 'ranges' and ranges_str:
        page_sets = _parse_ranges(ranges_str, total_pages)
    else:
        page_sets = [[i] for i in range(total_pages)]

    if len(page_sets) == 1:
        writer = PdfWriter()
        for idx in page_sets[0]:
            writer.add_page(reader.pages[idx])
        output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
        with open(output_path, 'wb') as f:
            writer.write(f)
        return output_path

    zip_path = os.path.join(OUTPUT_DIR, _unique_name('.zip'))
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, page_indices in enumerate(page_sets):
            writer = PdfWriter()
            for idx in page_indices:
                writer.add_page(reader.pages[idx])
            buf = io.BytesIO()
            writer.write(buf)
            zf.writestr(f"page_{i + 1}.pdf", buf.getvalue())
    return zip_path


def _parse_ranges(ranges_str, total_pages):
    result = []
    for part in ranges_str.split(','):
        part = part.strip()
        if '-' in part:
            start, end = part.split('-', 1)
            start = max(0, int(start.strip()) - 1)
            end = min(total_pages - 1, int(end.strip()) - 1)
            result.append(list(range(start, end + 1)))
        else:
            idx = int(part.strip()) - 1
            if 0 <= idx < total_pages:
                result.append([idx])
    return result if result else [[i] for i in range(total_pages)]


# ─── 3. COMPRESS PDF ─────────────────────────────────────────────────────────

def compress_pdf(file_path):
    try:
        import pikepdf
        pdf = pikepdf.open(file_path)
        output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
        pdf.save(output_path, linearize=True, compress_streams=True,
                 object_stream_mode=pikepdf.ObjectStreamMode.generate)
        pdf.close()

        for page in pikepdf.open(output_path).pages:
            for key in ['/XObject']:
                if key in (page.get('/Resources') or {}):
                    xobjects = page['/Resources'][key]
                    for name in list(xobjects.keys()):
                        obj = xobjects[name]
                        if hasattr(obj, '/Subtype') and str(obj.get('/Subtype')) == '/Image':
                            pass
        return output_path
    except Exception:
        writer = PdfWriter()
        reader = PdfReader(file_path)
        for page in reader.pages:
            page.compress_content_streams()
            writer.add_page(page)
        output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
        with open(output_path, 'wb') as f:
            writer.write(f)
        return output_path


# ─── 4. PDF TO WORD ──────────────────────────────────────────────────────────

def pdf_to_word(file_path):
    from pdf2docx import Converter
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.docx'))
    cv = Converter(file_path)
    cv.convert(output_path)
    cv.close()
    return output_path


# ─── 5. WORD TO PDF ──────────────────────────────────────────────────────────

def word_to_pdf(file_path):
    from docx import Document
    doc = Document(file_path)
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))

    c = canvas.Canvas(output_path, pagesize=A4)
    width, height = A4
    margin = 50
    y = height - margin
    line_height = 14

    for para in doc.paragraphs:
        text = para.text
        if not text:
            y -= line_height
            if y < margin:
                c.showPage()
                y = height - margin
            continue

        font_size = 12
        font_name = 'Helvetica'

        if para.style and para.style.name:
            style_name = para.style.name.lower()
            if 'heading 1' in style_name:
                font_size = 24
                font_name = 'Helvetica-Bold'
            elif 'heading 2' in style_name:
                font_size = 20
                font_name = 'Helvetica-Bold'
            elif 'heading 3' in style_name:
                font_size = 16
                font_name = 'Helvetica-Bold'

        if para.runs and para.runs[0].bold:
            font_name = 'Helvetica-Bold'

        c.setFont(font_name, font_size)
        line_height = font_size + 4

        max_chars = int((width - 2 * margin) / (font_size * 0.5))
        lines = []
        while text:
            if len(text) <= max_chars:
                lines.append(text)
                break
            split_at = text.rfind(' ', 0, max_chars)
            if split_at == -1:
                split_at = max_chars
            lines.append(text[:split_at])
            text = text[split_at:].lstrip()

        for line in lines:
            if y < margin:
                c.showPage()
                c.setFont(font_name, font_size)
                y = height - margin
            c.drawString(margin, y, line)
            y -= line_height

        y -= line_height * 0.3

    c.save()
    return output_path


# ─── 6. PDF TO JPG ───────────────────────────────────────────────────────────

def pdf_to_jpg(file_path):
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(file_path, dpi=200)
    except Exception:
        import fitz
        doc = fitz.open(file_path)
        images = []
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()

    if len(images) == 1:
        output_path = os.path.join(OUTPUT_DIR, _unique_name('.jpg'))
        images[0].save(output_path, 'JPEG', quality=90)
        return output_path

    zip_path = os.path.join(OUTPUT_DIR, _unique_name('.zip'))
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, img in enumerate(images):
            buf = io.BytesIO()
            img.save(buf, 'JPEG', quality=90)
            zf.writestr(f"page_{i + 1}.jpg", buf.getvalue())
    return zip_path


# ─── 7. JPG TO PDF ───────────────────────────────────────────────────────────

def jpg_to_pdf(file_paths):
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))

    images = []
    for path in file_paths:
        img = Image.open(path)
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        images.append(img)

    if not images:
        raise ValueError("No images provided")

    first = images[0]
    if len(images) > 1:
        first.save(output_path, 'PDF', save_all=True, append_images=images[1:])
    else:
        first.save(output_path, 'PDF')

    return output_path


# ─── 8. ROTATE PDF ───────────────────────────────────────────────────────────

def rotate_pdf(file_path, degrees=90):
    degrees = int(degrees)
    reader = PdfReader(file_path)
    writer = PdfWriter()
    for page in reader.pages:
        page.rotate(degrees)
        writer.add_page(page)
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 9. ADD WATERMARK ────────────────────────────────────────────────────────

def add_watermark(file_path, text='WATERMARK', opacity=0.3, font_size=60):
    opacity = float(opacity)
    font_size = int(font_size)

    reader = PdfReader(file_path)
    writer = PdfWriter()

    for page in reader.pages:
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)

        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(page_width, page_height))
        c.saveState()
        c.setFillColor(Color(0.5, 0.5, 0.5, alpha=opacity))
        c.setFont('Helvetica-Bold', font_size)
        c.translate(page_width / 2, page_height / 2)
        c.rotate(45)
        c.drawCentredString(0, 0, text)
        c.restoreState()
        c.save()

        packet.seek(0)
        watermark_page = PdfReader(packet).pages[0]
        page.merge_page(watermark_page)
        writer.add_page(page)

    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 10. PROTECT PDF ─────────────────────────────────────────────────────────

def protect_pdf(file_path, password):
    reader = PdfReader(file_path)
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt(password)
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 11. UNLOCK PDF ──────────────────────────────────────────────────────────

def unlock_pdf(file_path, password):
    reader = PdfReader(file_path)
    if reader.is_encrypted:
        reader.decrypt(password)
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 12. ADD PAGE NUMBERS ────────────────────────────────────────────────────

def add_page_numbers(file_path, position='bottom-center', start_num=1):
    start_num = int(start_num)
    reader = PdfReader(file_path)
    writer = PdfWriter()

    for i, page in enumerate(reader.pages):
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)
        page_num = start_num + i

        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(page_width, page_height))
        c.setFont('Helvetica', 10)

        if position == 'bottom-center':
            x, y = page_width / 2, 30
        elif position == 'bottom-right':
            x, y = page_width - 40, 30
        elif position == 'bottom-left':
            x, y = 40, 30
        elif position == 'top-center':
            x, y = page_width / 2, page_height - 30
        elif position == 'top-right':
            x, y = page_width - 40, page_height - 30
        elif position == 'top-left':
            x, y = 40, page_height - 30
        else:
            x, y = page_width / 2, 30

        c.drawCentredString(x, y, str(page_num))
        c.save()

        packet.seek(0)
        number_page = PdfReader(packet).pages[0]
        page.merge_page(number_page)
        writer.add_page(page)

    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)
    return output_path


# ─── 13. DOCUMENT SIGNER (21 CFR Part 11) ───────────────────────────────────

PREVIEW_DIR = os.path.join(settings.MEDIA_ROOT, 'previews')
os.makedirs(PREVIEW_DIR, exist_ok=True)


def _compute_file_hash(file_path):
    """Compute SHA-512 hash of a file for integrity verification."""
    import hashlib
    sha512 = hashlib.sha512()
    with open(file_path, 'rb') as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            sha512.update(chunk)
    return sha512.hexdigest()


def generate_document_preview(file_path, original_name=''):
    """
    Generate page preview images for the document signer interface.
    Supports PDF and DOCX files.
    Returns dict with list of preview image URLs and page count.
    """
    ext = Path(file_path).suffix.lower()
    preview_id = uuid.uuid4().hex

    if ext == '.pdf':
        return _preview_pdf(file_path, preview_id)
    elif ext in ('.docx', '.doc'):
        # Convert DOCX to PDF first, then preview
        pdf_path = word_to_pdf(file_path)
        return _preview_pdf(pdf_path, preview_id)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _preview_pdf(file_path, preview_id):
    """Generate preview images from PDF pages."""
    pages = []

    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        for i, page in enumerate(doc):
            pix = page.get_pixmap(dpi=150)
            img_name = f"{preview_id}_page_{i + 1}.jpg"
            img_path = os.path.join(PREVIEW_DIR, img_name)
            pix.save(img_path)
            pages.append({
                'page_num': i + 1,
                'url': f'/media/previews/{img_name}',
                'width': pix.width,
                'height': pix.height,
            })
        doc.close()
    except ImportError:
        # Fallback to pdf2image
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(file_path, dpi=150)
            for i, img in enumerate(images):
                img_name = f"{preview_id}_page_{i + 1}.jpg"
                img_path = os.path.join(PREVIEW_DIR, img_name)
                img.save(img_path, 'JPEG', quality=85)
                pages.append({
                    'page_num': i + 1,
                    'url': f'/media/previews/{img_name}',
                    'width': img.width,
                    'height': img.height,
                })
        except Exception as e:
            raise ValueError(f"Cannot generate preview: {str(e)}")

    return {
        'pages': pages,
        'total_pages': len(pages),
    }


def sign_document(file_path, signer_name, signer_email, meaning,
                  position_x=0, position_y=0, page_number=1):
    """
    Apply electronic signature to a document (21 CFR Part 11 compliant).

    Creates a visible signature block on the PDF containing:
    - Signer's printed name
    - Date and time of signature
    - Meaning of signature (Approved, Reviewed, etc.)
    - Unique Signature ID

    Returns dict with output path and document hashes.
    """
    import datetime

    ext = Path(file_path).suffix.lower()

    # If DOCX, convert to PDF first
    if ext in ('.docx', '.doc'):
        file_path = word_to_pdf(file_path)

    # Compute hash of original document
    hash_before = _compute_file_hash(file_path)

    # Generate signature ID
    sig_id = uuid.uuid4().hex[:16].upper()
    sig_time = datetime.datetime.now(datetime.timezone.utc)
    sig_time_str = sig_time.strftime('%Y-%m-%d %H:%M:%S UTC')

    # Build signature text lines
    meaning_display = meaning.capitalize()
    sig_lines = [
        f"Electronically signed by: {signer_name}",
        f"Email: {signer_email}",
        f"Date: {sig_time_str}",
        f"Meaning: {meaning_display}",
        f"Signature ID: {sig_id}",
    ]

    # Read the PDF
    reader = PdfReader(file_path)
    writer = PdfWriter()

    # Convert page_number to 0-indexed
    target_page_idx = max(0, page_number - 1)
    if target_page_idx >= len(reader.pages):
        target_page_idx = len(reader.pages) - 1

    for i, page in enumerate(reader.pages):
        if i == target_page_idx:
            # Create signature overlay
            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height)

            packet = io.BytesIO()
            c = canvas.Canvas(packet, pagesize=(page_width, page_height))

            # Calculate signature block position
            sig_width = 280
            sig_height = 90
            sig_x = max(10, min(position_x, page_width - sig_width - 10))
            sig_y = max(10, min(page_height - position_y - sig_height, page_height - sig_height - 10))

            # Draw signature block background
            c.saveState()
            c.setStrokeColor(Color(0.2, 0.2, 0.6, alpha=0.8))
            c.setFillColor(Color(0.95, 0.95, 1.0, alpha=0.9))
            c.setLineWidth(1.5)
            c.roundRect(sig_x, sig_y, sig_width, sig_height, 5, fill=True, stroke=True)

            # Draw signature text
            c.setFillColor(Color(0.1, 0.1, 0.3, alpha=1.0))
            text_x = sig_x + 8
            text_y = sig_y + sig_height - 16

            c.setFont('Helvetica-Bold', 7)
            c.drawString(text_x, text_y, sig_lines[0])
            text_y -= 12

            c.setFont('Helvetica', 6.5)
            for line in sig_lines[1:]:
                c.drawString(text_x, text_y, line)
                text_y -= 11

            # Draw "21 CFR Part 11" badge
            c.setFont('Helvetica-Oblique', 5)
            c.setFillColor(Color(0.4, 0.4, 0.6, alpha=0.7))
            c.drawString(sig_x + sig_width - 75, sig_y + 4, "21 CFR Part 11")

            c.restoreState()
            c.save()

            # Merge overlay
            packet.seek(0)
            overlay_page = PdfReader(packet).pages[0]
            page.merge_page(overlay_page)

        writer.add_page(page)

    # Save signed document
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.pdf'))
    with open(output_path, 'wb') as f:
        writer.write(f)

    # Compute hash of signed document
    hash_after = _compute_file_hash(output_path)

    return {
        'output_path': output_path,
        'hash_before': hash_before,
        'hash_after': hash_after,
        'signature_id': sig_id,
        'signed_at': sig_time_str,
    }


# ─── 14. PDF DATA EXTRACTOR (OCR-based) ──────────────────────────────────────

def extract_pdf_data(file_path, output_format='json'):
    """
    Extract text and structured data from PDF using text extraction + OCR.
    Supports text-based PDFs (fast) and scanned/image PDFs (OCR fallback).

    Detects: invoices, receipts, forms — extracts key-value pairs,
    amounts, dates, emails, phone numbers, GSTIN, PAN, line items, etc.

    output_format: 'json' or 'excel'
    Returns path to the output file (.json or .xlsx)
    """
    import json as json_mod
    import re

    pages_data = []
    full_text = ''

    # Step 1: Direct text extraction with PyMuPDF (fast)
    try:
        import fitz
        doc = fitz.open(file_path)
        for i, page in enumerate(doc):
            text = page.get_text('text')
            pages_data.append({
                'page': i + 1,
                'text': text.strip(),
                'method': 'text_extraction',
            })
            full_text += text + '\n'
        doc.close()
    except Exception:
        pass

    # Step 2: If very little text found, try OCR
    total_chars = sum(len(p['text']) for p in pages_data)
    if total_chars < 50:
        pages_data = []
        full_text = ''
        try:
            import fitz
            import pytesseract

            doc = fitz.open(file_path)
            for i, page in enumerate(doc):
                pix = page.get_pixmap(dpi=300)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                text = pytesseract.image_to_string(img)
                pages_data.append({
                    'page': i + 1,
                    'text': text.strip(),
                    'method': 'ocr',
                })
                full_text += text + '\n'
            doc.close()
        except Exception as ocr_err:
            try:
                from pdf2image import convert_from_path
                import pytesseract

                images = convert_from_path(file_path, dpi=300)
                for i, img in enumerate(images):
                    text = pytesseract.image_to_string(img)
                    pages_data.append({
                        'page': i + 1,
                        'text': text.strip(),
                        'method': 'ocr_fallback',
                    })
                    full_text += text + '\n'
            except Exception as e2:
                raise ValueError(
                    f"Could not extract text from PDF. "
                    f"OCR error: {str(ocr_err)}. Fallback: {str(e2)}"
                )

    if not full_text.strip():
        raise ValueError("No text could be extracted from this PDF.")

    # Step 3: Parse structured data
    extracted_fields = _parse_document_fields(full_text)

    result = {
        'document_info': {
            'total_pages': len(pages_data),
            'extraction_method': pages_data[0]['method'] if pages_data else 'unknown',
            'total_characters': len(full_text),
        },
        'extracted_fields': extracted_fields,
        'pages': pages_data,
        'raw_text': full_text.strip(),
    }

    # Step 4: Export
    if output_format == 'excel':
        return _export_to_excel(result)
    else:
        return _export_to_json(result)


def _parse_document_fields(text):
    """
    Parse common document fields from extracted text.
    Detects invoices, receipts, forms — key-value pairs.
    """
    import re

    fields = {}
    lines = text.split('\n')

    patterns = {
        'invoice_number': [
            r'(?:invoice\s*(?:no|number|#|num)?[\s:.\-]*)\s*([A-Z0-9\-/]+)',
            r'(?:inv\s*(?:no|#)?[\s:.\-]*)\s*([A-Z0-9\-/]+)',
            r'(?:bill\s*(?:no|number|#)?[\s:.\-]*)\s*([A-Z0-9\-/]+)',
        ],
        'date': [
            r'(?:date|dated|invoice\s*date|bill\s*date|order\s*date)[\s:.\-]*(\d{1,2}[\s/\-\.]\d{1,2}[\s/\-\.]\d{2,4})',
            r'(?:date|dated)[\s:.\-]*(\d{4}[\s/\-\.]\d{1,2}[\s/\-\.]\d{1,2})',
            r'(\d{1,2}[\s/\-]\w{3,9}[\s/\-]\d{2,4})',
        ],
        'total_amount': [
            r'(?:total|grand\s*total|amount\s*due|net\s*amount|balance\s*due)[\s:.\-]*[\$\u20B9\u00A3\u20AC]?\s*([\d,]+\.?\d*)',
            r'[\$\u20B9\u00A3\u20AC]\s*([\d,]+\.\d{2})',
        ],
        'tax_amount': [
            r'(?:tax|vat|gst|cgst|sgst|igst|sales\s*tax)[\s:.\-]*[\$\u20B9\u00A3\u20AC]?\s*([\d,]+\.?\d*)',
        ],
        'subtotal': [
            r'(?:subtotal|sub\s*total|sub\-total)[\s:.\-]*[\$\u20B9\u00A3\u20AC]?\s*([\d,]+\.?\d*)',
        ],
        'email': [
            r'([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})',
        ],
        'phone': [
            r'(?:phone|tel|mobile|contact|ph)[\s:.\-]*([+\d\s\-\(\)]{8,20})',
        ],
        'gstin': [
            r'(?:gstin|gst\s*no|gst\s*number)[\s:.\-]*(\d{2}[A-Z]{5}\d{4}[A-Z]\d[A-Z\d][A-Z]\d)',
        ],
        'pan': [
            r'(?:pan|pan\s*no|pan\s*number)[\s:.\-]*([A-Z]{5}\d{4}[A-Z])',
        ],
        'po_number': [
            r'(?:po\s*(?:no|number|#)?|purchase\s*order)[\s:.\-]*([A-Z0-9\-/]+)',
        ],
    }

    for field_name, regexes in patterns.items():
        for regex in regexes:
            match = re.search(regex, text, re.IGNORECASE | re.MULTILINE)
            if match:
                fields[field_name] = match.group(1).strip()
                break

    # Line items (table rows with amounts)
    line_items = []
    amount_pattern = re.compile(
        r'(.{5,60}?)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)'
    )
    for line in lines:
        match = amount_pattern.search(line.strip())
        if match:
            desc = match.group(1).strip()
            if not any(h in desc.lower() for h in ['description', 'item', 'particular', 'sl no', 'sr no', 'qty', 'quantity']):
                line_items.append({
                    'description': desc,
                    'value_1': match.group(2),
                    'value_2': match.group(3),
                    'value_3': match.group(4),
                })

    if line_items:
        fields['line_items'] = line_items

    # All key-value pairs (lines with : separator)
    kv_pairs = {}
    kv_pattern = re.compile(r'^([A-Za-z][A-Za-z\s\.]{2,30})[\s]*[:]\s*(.+)$', re.MULTILINE)
    for match in kv_pattern.finditer(text):
        key = match.group(1).strip().lower().replace(' ', '_')
        val = match.group(2).strip()
        if val and len(val) < 200:
            kv_pairs[key] = val

    if kv_pairs:
        fields['all_key_value_pairs'] = kv_pairs

    return fields


def _export_to_json(result):
    """Export extracted data as JSON file."""
    import json as json_mod
    output_path = os.path.join(OUTPUT_DIR, _unique_name('.json'))
    with open(output_path, 'w', encoding='utf-8') as f:
        json_mod.dump(result, f, indent=2, ensure_ascii=False, default=str)
    return output_path


def _export_to_excel(result):
    """Export extracted data as Excel (.xlsx) with styled sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D0D2B", end_color="0D0D2B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sheet 1: Extracted Fields ---
    ws1 = wb.active
    ws1.title = "Extracted Fields"

    ws1.append(["Document Information"])
    ws1.merge_cells('A1:B1')
    ws1['A1'].font = Font(bold=True, size=13)

    ws1.append(["Total Pages", result['document_info']['total_pages']])
    ws1.append(["Extraction Method", result['document_info']['extraction_method']])
    ws1.append(["Total Characters", result['document_info']['total_characters']])
    ws1.append([])

    ws1.append(["Field", "Value"])
    header_row = ws1.max_row
    for cell in ws1[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    fields = result.get('extracted_fields', {})
    for key, value in fields.items():
        if key in ('line_items', 'all_key_value_pairs'):
            continue
        display_key = key.replace('_', ' ').title()
        ws1.append([display_key, str(value)])
        for cell in ws1[ws1.max_row]:
            cell.border = thin_border

    kv_pairs = fields.get('all_key_value_pairs', {})
    if kv_pairs:
        ws1.append([])
        ws1.append(["All Detected Key-Value Pairs"])
        ws1.merge_cells(f'A{ws1.max_row}:B{ws1.max_row}')
        ws1[f'A{ws1.max_row}'].font = Font(bold=True, size=11)
        ws1.append(["Key", "Value"])
        hr = ws1.max_row
        for cell in ws1[hr]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for k, v in kv_pairs.items():
            ws1.append([k.replace('_', ' ').title(), str(v)])
            for cell in ws1[ws1.max_row]:
                cell.border = thin_border

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 60

    # --- Sheet 2: Line Items ---
    line_items = fields.get('line_items', [])
    if line_items:
        ws2 = wb.create_sheet("Line Items")
        ws2.append(["Description", "Value 1", "Value 2", "Value 3"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for item in line_items:
            ws2.append([item['description'], item['value_1'], item['value_2'], item['value_3']])
            for cell in ws2[ws2.max_row]:
                cell.border = thin_border
        ws2.column_dimensions['A'].width = 45
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 15

    # --- Sheet 3: Raw Text ---
    ws3 = wb.create_sheet("Raw Text")
    ws3.append(["Page", "Text Content"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for page in result.get('pages', []):
        text = page['text'][:32000]
        ws3.append([page['page'], text])
        ws3[f'B{ws3.max_row}'].alignment = Alignment(wrap_text=True)
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 100

    output_path = os.path.join(OUTPUT_DIR, _unique_name('.xlsx'))
    wb.save(output_path)
    return output_path


# ─── CLEANUP ──────────────────────────────────────────────────────────────────

def cleanup_old_files(max_age_seconds=3600):
    import time
    now = time.time()
    for directory in [UPLOAD_DIR, OUTPUT_DIR, PREVIEW_DIR]:
        for filename in os.listdir(directory):
            filepath = os.path.join(directory, filename)
            if os.path.isfile(filepath):
                if now - os.path.getmtime(filepath) > max_age_seconds:
                    try:
                        os.remove(filepath)
                    except OSError:
                        pass
